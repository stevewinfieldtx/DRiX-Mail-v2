import io
from fastapi import UploadFile
from openpyxl import Workbook
from app.main import _tabular,_suggest

def test_csv_real_headers_are_inspected_and_suggested():
    file=UploadFile(filename="leads.csv",file=io.BytesIO(b"Organization,Website,Decision Maker,Work Email\nAcme,https://acme.test,Ada,ada@acme.test\n"))
    headers,rows=_tabular(file)
    mapping=_suggest(headers)
    assert mapping=={"company_url":"Website","company_name":"Organization","contact_name":"Decision Maker","contact_email":"Work Email"}
    assert rows[0]["Organization"]=="Acme"

def test_xlsx_headers_and_rows_are_read():
    book=Workbook();sheet=book.active
    sheet.append(["Company","Domain","Job Title"])
    sheet.append(["Beta","https://beta.test","CTO"])
    data=io.BytesIO();book.save(data);data.seek(0)
    headers,rows=_tabular(UploadFile(filename="leads.xlsx",file=data))
    assert headers==["Company","Domain","Job Title"]
    assert rows==[{"Company":"Beta","Domain":"https://beta.test","Job Title":"CTO"}]
    assert _suggest(headers)["company_name"]=="Company"

def test_xlsx_report_title_is_not_treated_as_column_header():
    book=Workbook();sheet=book.active
    sheet.merge_cells("A1:D1");sheet["A1"]="fram^ — Consolidated Warm-Contact List"
    sheet.append([])
    sheet.append(["Company Name","URL","Contact Name","Email"])
    sheet.append(["Acme","https://acme.test","Ada","ada@acme.test"])
    data=io.BytesIO();book.save(data);data.seek(0)
    headers,rows=_tabular(UploadFile(filename="warm-list.xlsx",file=data))
    assert headers==["Company Name","URL","Contact Name","Email"]
    assert rows[0]["URL"]=="https://acme.test"
    assert _suggest(headers)["company_url"]=="URL"

def test_xlsx_blank_leading_rows_are_ignored():
    book=Workbook();sheet=book.active
    sheet.append([]);sheet.append([])
    sheet.append(["Organization","Website"])
    sheet.append(["Beta","https://beta.test"])
    data=io.BytesIO();book.save(data);data.seek(0)
    headers,rows=_tabular(UploadFile(filename="leads.xlsx",file=data))
    assert headers==["Organization","Website"]
    assert len(rows)==1
def test_empty_xlsx_is_safe():
    book=Workbook();book.active.delete_rows(1,10)
    data=io.BytesIO();book.save(data);data.seek(0)
    headers,rows=_tabular(UploadFile(filename="empty.xlsx",file=data))
    assert rows==[]
